import time

import openpyxl

from Config.log import logger


def run_save_to_xls(subdomain_list, port_list, filename):
    """
    将扫描结果保存到excel
    :return:
    """
    # 创建excel
    wb = openpyxl.Workbook()
    # 创建sheet
    ws_subdomain = wb.create_sheet(title='子域名')
    ws_port = wb.create_sheet(title='端口')
    wb.remove(wb['Sheet'])
    # 写入子域名
    save_subdomain(ws_subdomain, subdomain_list)
    # 写入端口
    save_port(ws_port, port_list)
    # 保存文件
    if filename:
        save_file_name = './results/' + filename
    else:
        save_file_name = './results/' + f'{time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())}_result.xlsx'
    wb.save(save_file_name)
    logger.log('ALERT', f'将子域名保存到文件: {save_file_name}')
    wb.close()
    return save_file_name


def save_subdomain(ws, subdomain):
    """
    将子域名写入excel
    :param ws:
    :param subdomain:
    :return:
    """
    if subdomain:
        ws.cell(row=1, column=1).value = '子域名'
        num = 1
        for i in subdomain:
            ws.cell(row=num + 1, column=1).value = i
            num += 1
    else:
        logger.log('DEBUG', '没有收集到任何子域名')
        pass


def save_port(ws, port):
    """
    将端口写入excel
    :param ws:
    :param port:
    :return:
    """
    if port:
        ws['A1'] = 'IP'
        ws['B1'] = '端口'
        ws['C1'] = '协议'
        num = 1
        for i in port:
            ws.cell(row=num + 1, column=1).value = i['IPAddr']
            ws.cell(row=num + 1, column=2).value = i['Port']
            ws.cell(row=num + 1, column=3).value = i['Protocol']
            num += 1
    else:
        logger.log('DEBUG', '没有收集到任何端口')
        pass
